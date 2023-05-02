from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
#import requests
from openpyxl import load_workbook
from datetime import datetime
import datetime
from datetime import timedelta
import os
from io import BytesIO
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import pyautogui
import win32clipboard
from PIL import Image
import pandas as pd
import xlwings  as xw
import numpy as np
import cv2
import subprocess
import schedule
import math
import chromedriver_autoinstaller

chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
driver_path = f'./{chrome_ver}/chromedriver.exe'
if os.path.exists(driver_path):
    print(f"chrom driver is insatlled: {driver_path}")
else:
    print(f"install the chrome driver(ver: {chrome_ver})")
    chromedriver_autoinstaller.install(True)
#chromedriver_autoinstaller.install()
data=MIMEMultipart()

print(cv2)
dataDict={0:'(월)', 1:'(화)', 2:'(수)', 3:'(목)', 4:'(금)', 5:'(토)', 6:'(일)'} #요일 출력 방식 지정

now = datetime.datetime.now().strftime("%H:%M")

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches',['enable-logging'])
options.add_experimental_option("prefs",{
    "download.default_directory": os.getcwd(),
  "download.prompt_for_download": False,
  "download.directory_upgrade": True,
  "safebrowsing.enabled": True
})

wb1 = load_workbook(filename=os.getcwd()+'/로그인.xlsx',data_only=True)  
# 엑셀파일로 아이디 비번 입력받으려면 해당 주석 활성화
ws1 = wb1[wb1.sheetnames[0]]

l_id = ws1.cell(row=1,column=2).value #아이디 이메일 주소까지 전부 입력
l_pw = ws1.cell(row=2,column=2).value #비밀번호 

def start_id() :                                    
    print("아이디를 입력해주세요")
    id = input()
    if "@3frame.com" in id :
        print("아이디가 입력되었습니다.")
        return(id)
    else :
        print("아이디를 확인해주세요")
        start_id()


def com(driver):                   #지각자 명단 엑셀의 데이터값이 1명 이상일경우 메일 전송으로 1명 미만일경우 진행 안함.
    wb2 = load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/지각자명단.xlsx',data_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    lis = len(pd.DataFrame(ws2))
    
    if lis >= 2:

        sendmail(ws2,driver)
       

    else:
        
        return print("완료")


def sendmail(ws2, driver):

    global data
    weekdays = datetime.datetime.today().weekday()
    if weekdays<4:
        day_w = datetime.datetime.today() + timedelta(days=1) #발생일이 목요일이면 다음날을 완료기한으로 도출
    if weekdays>=4:
        day_w = datetime.datetime.today() + timedelta(days=3) #발생일이 금요일이면 월요일을 완료기한으로 도출

    p_day_w=day_w.strftime("%Y.%m.%d")+dataDict[day_w.weekday()] #완료기한 년.월.일(요일) 방식으로 저장
    r_len = len(pd.DataFrame(ws2))+1
    for 행 in ws2.iter_rows(min_row=2):
        t_day=행[2].value                            #엑셀 파일에서 일자 변수 지정
        tt_day=행[3].value                           #엑셀 파일에서 요일 변수 지정
        #tt_time=ws2.cell(row=l,column=5).value#.strftime("%H:%M")
        t_name=행[1].value                           #엑셀 파일에서 이름 변수 지정
        t_mail=행[7].value                           #엑셀 파일에서 메일 변수 지정
        t_cc=행[8].value                             #엑셀 파일에서 참조자 변수 지정

        if 행[4].value == "미출근" :                 #엑셀 파일상 출근시간이 명시되어 있지 않는경우 메일 내용 변경
            tt_time="출근 인증 없음"
            mail_data = """안녕하세요 경영기획실입니다. <br>
                <br>
                1) 요청사항 : """+t_day +""+tt_day+" "+str(tt_time)+"""으로 출근 등록 및 아래 증빙이 없을경우 부재일정 반반차(0.25) 상신.(22.10.04. 이후부터 적용)<br>
                2) 완료기한 : """+p_day_w+"""<br>
                <br>
               요청드린 기한까지 회신 부탁드리며, 증빙서류가 있는경우 제출바랍니다.<br>
                증빙서류 예시 : 당일 교통 카드 내역(도착지기준) 제출시 지각으로 간주하지 않음(단, 출근시간 10분전 내역에 한해서 소명가능)<b><br>
                예)9시 출근일 경우, 8:50 양재역 또는 회사 부근 하차내역 必. 인사담당자(경영기획-어유진 대리)에게 사유전달<br><br>
                </b>문의사항은 서포트 메일로 보내주시기 바랍니다.<br>
                <br>
                감사합니다.<br>"""

        else:
            tt_time=행[4].value.strftime("%H:%M")+" 출근 등록"       #엑셀 파일상 출근시간이 명시 되어 있는경우 기본 메일폼으로 진행
            mail_data = """안녕하세요 경영기획실입니다. <br>
                <br>
                1) 요청사항 : """+t_day +""+tt_day+" "+str(tt_time)+"""으로 증빙서류 제출 혹은 부재일정 반반차(0.25) 상신.<br>
                2) 완료기한 : """+p_day_w+"""<br>
                <br>
               요청드린 기한까지 회신 요청드리며,증빙서류가 있는경우 제출바랍니다.<br>
               증빙서류 예시 : 당일 지하철 지연증명서<br>
               인사담당자(경영기획-어유진 대리)에게 증빙서류 제출 및 전달<br>
                문의사항은 서포트 메일로 보내주시기 바랍니다.<br>
                <br>
                감사합니다.<br>"""
            
        #with open(os.getcwd()+'/'+time.strftime("%Y%m%d")+'/'+t_name+".png", 'rb') as fp:       #요일별 직원별 출근시간 이미지 가져오기
         #   img = MIMEImage(fp.read(), Name = "capture.png")
          #  img.add_header('Content-ID', '<capture>')
           # data.attach(img)


        

        def send_to_clipboard(clip_type, data):
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(clip_type, data)
            win32clipboard.CloseClipboard()

        filepath = os.getcwd() + "/"+time.strftime("%Y%m%d")+"/"+'/'+t_name+".png"
        image = Image.open(filepath)

        output = BytesIO()
        image.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()

        send_to_clipboard(win32clipboard.CF_DIB, data)

        driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
        time.sleep(3)
        #lis_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > table > tbody > tr:nth-child(1) > td > div > div.mInput_holder.parallel_first > div > input')
        #lis_btn.click()

        #send_btn = driver.find_element(By.CSS_SELECTOR,'#senderAddressList > li:nth-child(5) > a').click()

        elem_to = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(2) > td > div > div.mInput_holder > ul > li > div > div > input[type=text]').send_keys(t_mail)
        elem_cc = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(3) > td > div > ul > li > div > div > input[type=text]').send_keys(t_cc)
        bcc_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > table > tbody > tr:nth-child(3) > th > a').click()
        elem_bcc = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(4) > td > div > ul > li > div > div > input[type=text]').send_keys('support@3frame.com')
        impo_btn = driver.find_element(By.CSS_SELECTOR,'#impMail').click()
        elem_subject = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(5) > td > div > div > input').send_keys('안녕하세요 경영기획실입니다.')
        body_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
        elem_body = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_body > textarea:nth-child(1)')
        elem_body.send_keys(mail_data)
        html_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
        time.sleep(1)
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'v')
        
        send_btn = driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
        time.sleep(0.5)
        send_btn = driver.find_element(By.CSS_SELECTOR,'#modal-holder > div > div > div > div > button.lw_btn_point').click()
        time.sleep(5)
    return 


def cap_web(driver) :               #출근시간 캡쳐
    driver.get("https://3frame.ncpworkplace.com/tcs/manager/commute/list#")
   
    time.sleep(2)
    driver.maximize_window()
    
    num = driver.find_element(By.CSS_SELECTOR,'#totalCntText')

    table = driver.find_element(By.CSS_SELECTOR,'#timeDayList')
    tbody =table.find_element(By.CSS_SELECTOR,"#timeDayList > tbody")
    rows = tbody.find_elements(By.TAG_NAME,"tr")
    
    #for index, value in enumerate(rows):
     #   body=value.find_element(By.TAG_NAME,"td")[0]
      #  print(body.text)

    down_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div[2]/button[2]')
    down_btn.click()
    tbodys = table.find_element(By.CSS_SELECTOR,"#timeDayList > tbody")
    rowss= tbody.find_elements(By.TAG_NAME,"tr")
    n=driver.find_element(By.CSS_SELECTOR,'#totalCntText')
    
    p = math.ceil(int(driver.find_element(By.CSS_SELECTOR,'#totalCntText').text)/int(len(rowss)))

    #while driver.find_element_by_xpath('//*[@id="pageNav"]/li['+p+']') in str :

    

    if p > 1 :
        
        for p_lis in range(4,int(p)+3) :
            tbody = table.find_element(By.CSS_SELECTOR,"#timeDayList > tbody")
            rows= tbody.find_elements(By.TAG_NAME,"tr")
            lis_cap(rows,driver)
            driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div[3]/div[2]/li['+str(p_lis)+']').click()
            time.sleep(0.5)
        tbody = table.find_element(By.CSS_SELECTOR,"#timeDayList > tbody")
        rows= tbody.find_elements(By.TAG_NAME,"tr")
        lis_cap(rows,driver) 

    else :
        lis_cap(rows,driver)

def lis_cap(rows,driver):
    for i in range(2, len(rows)+1):
        i_str = str(i)
        name = driver.find_element(By.CSS_SELECTOR,'#timeDayList > tbody > tr:nth-child(' + i_str + ') > td:nth-child(3) > div > span > a')
        name_str = str(name.text)
        img = driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/div[2]/div/div[2]/table/tbody/tr['+ i_str +']')
        img_png = img.screenshot_as_png

        p_scroll = str(i + 5)

        with open(os.getcwd() + "/"+time.strftime("%Y%m%d")+"/" + name_str +".png","wb") as file :
            file.write(img_png)

        driver.execute_script("window.scrollTo(0,"+p_scroll+")")
        img_op = Image.open(os.getcwd()+'/'+time.strftime("%Y%m%d")+"/" + name_str +".png")

        img_resize = img_op.resize((1580,40),Image.Resampling.LANCZOS)
        cropimg=img_resize.crop((0,0,1304,40))
        cropimg.save(os.getcwd()+'/'+time.strftime("%Y%m%d") + "/" + name_str +".png")

        main_img = (os.path.dirname(os.path.realpath(__file__))+"/main.png")
        #main_img_re = main_img.resize((1865,76),Image.ANTIALIAS)
        d_img = (os.getcwd()+'/'+time.strftime("%Y%m%d") + "/" + name_str +".png")
        mimg_add = np.fromfile(main_img,np.uint8)
        dimg_add = np.fromfile(d_img,np.uint8)

        d_img2 =cv2.imdecode(dimg_add,cv2.IMREAD_COLOR)
        main_img2 = cv2.imdecode(mimg_add,cv2.IMREAD_COLOR)

        addimg=np.vstack((main_img2,d_img2))

        saving = os.getcwd()+'/'+time.strftime("%Y%m%d")+"/" + name_str +".png"
        result,encode_img = cv2.imencode(saving,addimg)
        time.sleep(0.5)
        if result :
            with open(saving,mode='w+b') as f:
                encode_img.tofile(f)    
    

def excel_run():
    path = os.getcwd()
    wb = xw.Book(path+'/테스트.xlsm')

    macro = wb.macro('Macro1')
    time.sleep(2)
    macro()

    
def rest_day(driver) :
    time.sleep(3)
    driver.get("https://3frame.ncpworkplace.com/hrs/manager/absence/timeOffList")
    time.sleep(3)
    lis_btn = driver.find_element(By.CSS_SELECTOR,'#fromYmd')
    lis_btn.click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('DELETE')
    time.sleep(3)

    lis_btn.send_keys(time.strftime("%Y%m%d"))

    ser_btn = driver.find_element(By.CSS_SELECTOR,'#searchForm > div > div > div.pull-left.ml-10 > button')
    ser_btn.click()


    time.sleep(2)
    d_btn = driver.find_element(By.CSS_SELECTOR,'body > div > div > div.flexible-content > div.content-body > div.form-row.row > button:nth-child(2)')
    d_btn.click()


def login():
    
    driver = webdriver.Chrome(options=options)
    driver.get("https://3frame.ncpworkplace.com/")


    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
    elem_id = driver.find_element(By.ID, 'user')
    elem_id.send_keys(l_id)

    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
    elem_pw = driver.find_element(By.ID, 'password')
    elem_pw.send_keys(l_pw)

    login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
    login_btn.click()


    time.sleep(5)

    cap_web(driver)
    rest_day(driver)
    excel_run()
    time.sleep(2)
    os.remove(os.getcwd() +'/commuteList.xlsx')
    os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
    
   

    com(driver)
    
def login_test():
    driver = webdriver.Chrome(options=options)
    driver.get("https://3frame.ncpworkplace.com/")


    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
    elem_id = driver.find_element(By.ID, 'user')
    elem_id.send_keys(l_id)

    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
    elem_pw = driver.find_element(By.ID, 'password')
    elem_pw.send_keys(l_pw)

    login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
    login_btn.click()


    time.sleep(5)

        
   

    com(driver)    

def time_down():
    driver = webdriver.Chrome(options=options)
    driver.get("https://3frame.ncpworkplace.com/")

    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
    elem_id = driver.find_element(By.ID, 'user')
    elem_id.send_keys(l_id)

    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
    elem_pw = driver.find_element(By.ID, 'password')
    elem_pw.send_keys(l_pw)

    login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
    login_btn.click()

    time.sleep(2)
    driver.get("https://3frame.ncpworkplace.com/tcs/manager/commute/list#")
    #url = "https://3frame.ncpworkplace.com/tcs/manager/commute/list#"

    down_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div[2]/button[2]')
    down_btn.click()
    time.sleep(2)


    

    def resting_day() :
        time.sleep(3)
        driver.get("https://3frame.ncpworkplace.com/hrs/manager/absence/timeOffList")
        time.sleep(3)
        lis_btn = driver.find_element(By.CSS_SELECTOR,'#fromYmd')
        time.sleep(3)
        lis_btn.click()
        time.sleep(3)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('DELETE')
        time.sleep(3)

        lis_btn.send_keys(time.strftime("%Y%m%d"))

        ser_btn = driver.find_element(By.CSS_SELECTOR,'#searchForm > div > div > div.pull-left.ml-10 > button')
        ser_btn.click()


        time.sleep(2)
        d_btn = driver.find_element(By.CSS_SELECTOR,'body > div > div > div.flexible-content > div.content-body > div.form-row.row > button:nth-child(2)')
        d_btn.click()
    
    resting_day()

    def sending ():
        for l in range(2,len(pd.DataFrame(ws3))+1):
            send_mail=ws3.cell(row=l,column=8).value
            t_cc=ws3.cell(row=l,column=9).value
            print(now + "2:00")

            mail_data = """안녕하세요 경영기획실입니다. <br>
              #  <br>
               # 1) 요청사항 : 출근미등록 으로  출근등록.<br>
               # 2) 완료기한 : """+datetime.datetime.today().strftime("%Y.%m.%d")+" "+(datetime.datetime.now() + datetime.timedelta(hours=1)).strftime("%H:%M")+"""<br>
               # <br>
               # 요청드린 기한까지 등록요청 요청드리며,<br>
               # 문의사항은 서포트 메일로 보내주시기 바랍니다.<br>
               # <br>
               # 감사합니다.<br>"""
        
            driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
            time.sleep(3)
            lis_btn = driver.find_element(By.CSS_SELECTOR,'#senderAddressSuggestLayerShowLnk')
            lis_btn.click()
            send_btn = driver.find_element(By.CSS_SELECTOR,'#senderAddressList > li:nth-child(5) > a').click()

            elem_to = driver.find_element(By.ID, 'toInput').send_keys(send_mail)
            elem_cc = driver.find_element(By.ID, 'ccInput').send_keys(t_cc)
            bcc_btn = driver.find_element(By.CSS_SELECTOR,'#icon_bccview').click()
            elem_bcc = driver.find_element(By.ID, 'bccInput').send_keys('support@3frame.com')
            impo_btn = driver.find_element(By.CSS_SELECTOR,'#priority').click()
            elem_subject = driver.find_element(By.ID, 'subject').send_keys('안녕하세요. 경영기획실입니다.')
            body_btn = driver.find_element(By.CSS_SELECTOR,'#divWrite > div.editorFrame > div.workseditor-classic > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
            elem_body = driver.find_element(By.CSS_SELECTOR,'#divWrite > div.editorFrame > div.workseditor-classic > div.editor_body > textarea:nth-child(1)')
            elem_body.send_keys(mail_data)
            html_btn = driver.find_element(By.CSS_SELECTOR,'#divWrite > div.editorFrame > div.workseditor-classic > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
            send_btn = driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
            time.sleep(2)

    if now > "10:05" and now <= "11:00":

        path = os.getcwd()
        wb = xw.Book(path+'/출근시간정리.xlsm')
        macro = wb.macro('Macro1')
        time.sleep(2)
        macro()
        time.sleep(2)
        wb3 = load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/근무시간-C.xlsx',data_only=True)
        ws3 = wb3[wb3.sheetnames[0]]
        
        list = len(pd.DataFrame(ws3))

        if list >= 2:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
            sending()
    
        else:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
    

    if now > "09:05" and now <= "10:00":

        path = os.getcwd()
        wb = xw.Book(path+'/출근시간정리.xlsm')
        macro = wb.macro('Macro1')
        time.sleep(2)
        macro()
        time.sleep(2)
        wb3 = load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/근무시간-B.xlsx',data_only=True)
        ws3 = wb3[wb3.sheetnames[0]]
        
        list = len(pd.DataFrame(ws3))

        if list >= 2:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
            sending()
    
        else:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
    
    #if lis >= 2 :

    if now > "08:05" and now <= "09:00":

        path = os.getcwd()
        wb = xw.Book(path+'/출근시간정리.xlsm')
        macro = wb.macro('Macro1')
        time.sleep(2)
        macro()
        time.sleep(2)
        wb3 = load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/근무시간-A.xlsx',data_only=True)
        ws3 = wb3[wb3.sheetnames[0]]
        
        list = len(pd.DataFrame(ws3))

        if list >= 2:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
            sending()
    
        else:
            time.sleep(2)
            os.remove(os.getcwd() +'/commuteList.xlsx')
            os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
        
        
    else:
        time.sleep(2)
       # os.remove(os.getcwd() +'/commuteList.xlsx')
       # os.remove(os.getcwd() +'/absenceTimeOffList.xlsx')
        print("완료")
        quit(driver)

def message():
    if weekdays2<5:
        print("스케쥴 실행중...")
        run()
    else :
        os.system('shutdown -s -f')
        print("휴일")

path = os.getcwd()
weekdays2 = datetime.datetime.today().weekday()

def createF(dir1):
    try:
        if not os.path.exists(dir1):
            os.makedirs(dir1)
    except OSError:
        print("폴더생성이 되지 않았습니다.")

def run():

    if time.strftime("%H:%M:%S") == "12:03:30" :
        print("현재시간 : "+time.strftime("%H:%M:%S"))
        login()
        
        

    if time.strftime("%H:%M:%S") == "15:19:00" :
        print("현재시간 : "+time.strftime("%H:%M:%S"))
        #login_test()
        #login()
                
    
    if time.strftime("%H:%M:%S") == "09:05:00" :
        print("현재시간 : "+time.strftime("%H:%M:%S"))
        #time_down()
        

    if time.strftime("%H:%M:%S") == "16:00:00" :
        if weekdays2<5:
        
            print("현재시간 : "+time.strftime("%H:%M:%S"))
        #time_down()
        else :
            os.system('shutdown -s -f')

    if time.strftime("%H:%M:%S") == "11:00:00" :
        cf = datetime.datetime.today()
        createF(os.getcwd()+"/"+str(cf.strftime("%Y%m%d")))


    else :
        print("현재시간 : "+time.strftime("%H:%M:%S"))
       # print("현재시간 : "+datetime.datetime.today().strftime("%Y.%m.%d"))
        
#l_id=start_id()                             #아이디 입력받기
#print("비밀번호를 입력해주세요")              #비밀번호 입력받기
#l_pw=input()

#if weekdays2<5:                             #평일에만 스케줄 실행
schedule.every(1).second.do(message)
    

#if weekdays2>=5:                            #주말에는 실행 안되도록 
   #schedule.every(1).minutes.do(print("휴일"))
 

while True:
    schedule.run_pending()
    time.sleep(1)

