import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtWidgets
import os
import time
import sys
from openpyxl import load_workbook
import xlwings as xw
import pandas as pd
import pythoncom
import datetime
from datetime import timedelta


database_wb=load_workbook(filename=os.getcwd()+'/사원명부.xlsm',data_only=True)
database_ws=database_wb['명단']
database_ws2=database_wb['발송여부']

class chrome_update():
    def ch_update():
        chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
        driver_path = f'./{chrome_ver}/chromedriver.exe'
        if os.path.exists(driver_path):
           
           pass
        else:
           
            chromedriver_autoinstaller.install(True)
        driver=webdriver.Chrome()
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        options.add_experimental_option("prefs",{
            "download.default_directory": os.getcwd(),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
        })
        #options.add_argument("headless")

class Db_update(QThread):
    label_msg=pyqtSignal(str)
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent
       
    def run(self):
        l_id = self.parent.id_insert.text()+"@3frame.com"
        l_pw = self.parent.pw_insert.text()
        chrome_update.ch_update
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        options.add_experimental_option("prefs",{
            "download.default_directory": os.getcwd(),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
            })
        options.add_argument("headless")
        driver = webdriver.Chrome(options=options)
        driver.get("https://3frame.ncpworkplace.com/")


        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
        elem_id = driver.find_element(By.ID, 'user')
        elem_id.send_keys(l_id)

        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
        elem_pw = driver.find_element(By.ID, 'password')
        elem_pw.send_keys(l_pw)

        login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
        login_btn.click()


        time.sleep(5)

        
        self.rest_day(driver)
        
        time.sleep(2)
        pythoncom.CoInitialize()
        app = xw.App(visible=False)
        wb2 = xw.Book(os.getcwd()+'/사원명부.xlsm')
        macro = wb2.macro('filter')
        macro()
        time.sleep(2)
        app.kill()
        self.label_msg.emit("DB up")
                        
    def rest_day(self,driver) :
        time.sleep(3)
        driver.get("https://3frame.ncpworkplace.com/hrs/manager/annualLeave/list")
        time.sleep(3)
        lis_btn = driver.find_element(By.CSS_SELECTOR,'#tap_area1 > div.form-row.row > button')
        lis_btn.click()

class send_mail(QThread):
    label_msg2=pyqtSignal(str)
    
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent

    def run(self) :
        weekdays = datetime.datetime.today().weekday()
        if weekdays<4:
            day_w = datetime.datetime.today() + timedelta(days=14) #발생일이 목요일이면 다음날을 완료기한으로 도출
        if weekdays>=4:
            day_w = datetime.datetime.today() + timedelta(days=17) #발생일이 금요일이면 월요일을 완료기한으로 도출
        if weekdays<4:
            day_w2 = datetime.datetime.today() + timedelta(days=7) #발생일이 목요일이면 다음날을 완료기한으로 도출
        if weekdays>=4:
            day_w2 = datetime.datetime.today() + timedelta(days=10) #발생일이 금요일이면 월요일을 완료기한으로 도출
        l_id = self.parent.id_insert.text()+"@3frame.com"
        l_pw = self.parent.pw_insert.text()
        chrome_update.ch_update
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        options.add_experimental_option("prefs",{
            "download.default_directory": os.getcwd(),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
            })
        options.add_argument("headless")
        driver = webdriver.Chrome(options=options)
        driver.get("https://3frame.ncpworkplace.com/")


        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
        elem_id = driver.find_element(By.ID, 'user')
        elem_id.send_keys(l_id)

        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
        elem_pw = driver.find_element(By.ID, 'password')
        elem_pw.send_keys(l_pw)

        login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
        login_btn.click()


        time.sleep(5)
        dataDict={0:'(월)', 1:'(화)', 2:'(수)', 3:'(목)', 4:'(금)', 5:'(토)', 6:'(일)'} #요일 출력 방식 지정
        time.sleep(2)
        
        p_day_w=day_w.strftime("%Y.%m.%d")+dataDict[day_w.weekday()] #완료기한 년.월.일(요일) 방식으로 저장
        end_time = day_w2.strftime("%Y.%m.%d")+dataDict[day_w2.weekday()] #완료기한 년.월.일(요일) 방식으로 저장
        try:
            dir_=os.getcwd()+"/"+str(time.strftime("%Y%m%d"))
            if os.path.exists(dir_):
                wb3=load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/발송용.xlsx',data_only=True)
                ws3_1 = wb3[wb3.sheetnames[0]]
                send_list = len(pd.DataFrame(ws3_1))
                ws3_2 = wb3[wb3.sheetnames[1]]
                re_send_list = len(pd.DataFrame(ws3_2))
                if send_list >1 and re_send_list>1:
                    for 행 in ws3_1.iter_rows(min_row=2):
                        s_mail = 행[2].value
                        s_name = 행[1].value
                        s_class = 행[3].value
                        s_restday = 행[5].value
                        s_useday = 행[6].value
                        s_leftday=행[7].value
                        mail_data = """<span style="font-family:Batang, 바탕, serif">
                                    <B>안녕하세요.<Br>
                                    경영기획실 입니다.<Br>
                                    미사용 연차에 대해 소멸 시점 전까지의 연차 사용계획서를 <Br>
                                        아래의 요청기한까지 작성하여 회신 요청드립니다.<Br>
                                        <table border="1">
                                        <th>부서</th>
                                        <th>이름</th>
                                        <th>발생연차</th>
                                        <th>사용연차</th>
                                        <th>미사용연차</th>
                                        <tr>
                                            <td>"""+str(s_class)+"""</td>
                                            <td>"""+str(s_name)+"""</td>
                                            <td>"""+str(s_restday)+"""</td>
                                            <td>"""+str(s_useday)+"""</td>
                                            <td style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);" data-origin-bgcolor="empty" bgcolor="rgb(249, 212, 0)">"""+str(s_leftday)+"""</td>
                                        </tr>
                                        </table>
                                    <Br>
                                    1. 요청일자 :<span style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);">"""+str(p_day_w)+""" 까지</span><Br>
                                    2. 작성예시 <Br>
                                    - 4/3(월) : 1일 사용<Br>
                                    - 5/10(수) : 0.5일 사용<Br>
                                    <Br>
                                    현재 입사일 기준으로 발생한 연(월)차를 발생일 기준 1년 내에 모두 소진해야 하는 것을 원칙으로 하고 있습니다.<Br>
                                    *시행일자 : 22.06.01<Br>
                                    위와 안내드린 바와 같이 사용 기간 내에 연차를 사용하지 않을 경우 자동 소멸 되오니,<Br>
                                    이점 참고하여 잔여 연차가 발생하지 않도록 계획하여 사용해 주시기 바랍니다.<Br>
                                    <Br>
                                    감사합니다.
                                    """
                        driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
                        time.sleep(3)
                        elem_to = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(2) > td > div > div.mInput_holder > ul > li > div > div > input[type=text]').send_keys(s_mail)
                        impo_btn = driver.find_element(By.CSS_SELECTOR,'#impMail').click()
                        elem_subject = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(5) > td > div > div > input').send_keys('[3FRAME] 미사용 연차 사용 계획서 요청')
                        body_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
                        elem_body = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_body > textarea:nth-child(1)')
                        elem_body.send_keys(mail_data)
                        html_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
                        time.sleep(1)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
                        time.sleep(0.5)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#modal-holder > div > div > div > div > button.lw_btn_point').click()
                        time.sleep(5)
                    for 행 in ws3_2.iter_rows(min_row=2):
                        s_mail = 행[2].value
                        s_name = 행[1].value
                        s_class = 행[3].value
                        s_restday = 행[5].value
                        s_useday = 행[6].value
                        s_leftday=행[7].value
                        mail_data = """<span style="font-family:Batang, 바탕, serif">
                                    <B>안녕하세요.<Br>
                                    경영기획실 입니다.<Br>
                                    미사용 연차에 대해 소멸 시점 전까지의 연차 사용계획서를 <Br>
                                        아래의 요청기한까지 작성하여 회신 요청드립니다.<Br>
                                        <table border="1">
                                        <th>부서</th>
                                        <th>이름</th>
                                        <th>발생연차</th>
                                        <th>사용연차</th>
                                        <th>미사용연차</th>
                                        <tr>
                                            <td>"""+str(s_class)+"""</td>
                                            <td>"""+str(s_name)+"""</td>
                                            <td>"""+str(s_restday)+"""</td>
                                            <td>"""+str(s_useday)+"""</td>
                                            <td style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);" data-origin-bgcolor="empty" bgcolor="rgb(249, 212, 0)">"""+str(s_leftday)+"""</td>
                                        </tr>
                                        </table>
                                    <Br>
                                    1. 요청일자 :<span style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);">"""+str(end_time)+""" 까지</span><Br>
                                    2. 작성예시 <Br>
                                    - 4/3(월) : 1일 사용<Br>
                                    - 5/10(수) : 0.5일 사용<Br>
                                    <Br>
                                    현재 입사일 기준으로 발생한 연(월)차를 발생일 기준 1년 내에 모두 소진해야 하는 것을 원칙으로 하고 있습니다.<Br>
                                    *시행일자 : 22.06.01<Br>
                                    위와 안내드린 바와 같이 사용 기간 내에 연차를 사용하지 않을 경우 자동 소멸 되오니,<Br>
                                    이점 참고하여 잔여 연차가 발생하지 않도록 계획하여 사용해 주시기 바랍니다.<Br>
                                    <Br>
                                    감사합니다.
                                    """
                        driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
                        time.sleep(3)
                        elem_to = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(2) > td > div > div.mInput_holder > ul > li > div > div > input[type=text]').send_keys(s_mail)
                        impo_btn = driver.find_element(By.CSS_SELECTOR,'#impMail').click()
                        elem_subject = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(5) > td > div > div > input').send_keys('[3FRAME] 미사용 연차 사용 계획서 요청')
                        body_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
                        elem_body = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_body > textarea:nth-child(1)')
                        elem_body.send_keys(mail_data)
                        html_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
                        time.sleep(1)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
                        time.sleep(0.5)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#modal-holder > div > div > div > div > button.lw_btn_point').click()
                        time.sleep(5)
                        
                elif send_list >1 and re_send_list == 1 :
                    for 행 in ws3_1.iter_rows(min_row=2):
                        s_mail = 행[2].value
                        s_name = 행[1].value
                        s_class = 행[3].value
                        s_restday = 행[5].value
                        s_useday = 행[6].value
                        s_leftday=행[7].value
                        mail_data = """<span style="font-family:Batang, 바탕, serif">
                                    <B>안녕하세요.<Br>
                                    경영기획실 입니다.<Br>
                                    미사용 연차에 대해 소멸 시점 전까지의 연차 사용계획서를 <Br>
                                        아래의 요청기한까지 작성하여 회신 요청드립니다.<Br>
                                        <table border="1">
                                        <th>부서</th>
                                        <th>이름</th>
                                        <th>발생연차</th>
                                        <th>사용연차</th>
                                        <th>미사용연차</th>
                                        <tr>
                                            <td>"""+str(s_class)+"""</td>
                                            <td>"""+str(s_name)+"""</td>
                                            <td>"""+str(s_restday)+"""</td>
                                            <td>"""+str(s_useday)+"""</td>
                                            <td style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);" data-origin-bgcolor="empty" bgcolor="rgb(249, 212, 0)">"""+str(s_leftday)+"""</td>
                                        </tr>
                                        </table>
                                    <Br>
                                    1. 요청일자 :<span style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);">"""+str(p_day_w)+""" 까지</span><Br>
                                    2. 작성예시 <Br>
                                    - 4/3(월) : 1일 사용<Br>
                                    - 5/10(수) : 0.5일 사용<Br>
                                    <Br>
                                    현재 입사일 기준으로 발생한 연(월)차를 발생일 기준 1년 내에 모두 소진해야 하는 것을 원칙으로 하고 있습니다.<Br>
                                    *시행일자 : 22.06.01<Br>
                                    위와 안내드린 바와 같이 사용 기간 내에 연차를 사용하지 않을 경우 자동 소멸 되오니,<Br>
                                    이점 참고하여 잔여 연차가 발생하지 않도록 계획하여 사용해 주시기 바랍니다.<Br>
                                    <Br>
                                    감사합니다.
                                    """
                        driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
                        time.sleep(3)
                        elem_to = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(2) > td > div > div.mInput_holder > ul > li > div > div > input[type=text]').send_keys(s_mail)
                        impo_btn = driver.find_element(By.CSS_SELECTOR,'#impMail').click()
                        elem_subject = driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(5) > td > div > div > input').send_keys('[3FRAME] 미사용 연차 사용 계획서 요청')
                        body_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
                        elem_body = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_body > textarea:nth-child(1)')
                        elem_body.send_keys(mail_data)
                        html_btn = driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
                        time.sleep(1)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
                        time.sleep(0.5)
                        send_btn = driver.find_element(By.CSS_SELECTOR,'#modal-holder > div > div > div > div > button.lw_btn_point').click()
                        time.sleep(5)
                        
                elif send_list ==1 and re_send_list >1 :
                   
                    for 행 in ws3_2.iter_rows(min_row=2):
                        driver.get("https://mail.worksmobile.com/write/popup/?orderType=new")
                        time.sleep(3)
                        s_mail = 행[2].value
                        s_name = 행[1].value
                        s_class = 행[3].value
                        s_restday = 행[5].value
                        s_useday = 행[6].value
                        s_leftday=행[7].value
                        
                        driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(2) > td > div > div.mInput_holder > ul > li > div > div > input[type=text]').send_keys(s_mail)
                        driver.find_element(By.CSS_SELECTOR,'#impMail').click()
                        driver.find_element(By.CSS_SELECTOR, '#contact_layer_offset_elem > table > tbody > tr:nth-child(5) > td > div > div > input').send_keys('[3FRAME] 미사용 연차 사용 계획서 요청')
                        driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(2)').click()
                        mail_data = """<span style="font-family:Batang, 바탕, serif">
                                    <B>안녕하세요.<Br>
                                    경영기획실 입니다.<Br>
                                    미사용 연차에 대해 소멸 시점 전까지의 연차 사용계획서를 <Br>
                                        아래의 요청기한까지 작성하여 회신 요청드립니다.<Br>
                                        <table border="1">
                                        <th>부서</th>
                                        <th>이름</th>
                                        <th>발생연차</th>
                                        <th>사용연차</th>
                                        <th>미사용연차</th>
                                        <tr>
                                            <td>"""+str(s_class)+"""</td>
                                            <td>"""+str(s_name)+"""</td>
                                            <td>"""+str(s_restday)+"""</td>
                                            <td>"""+str(s_useday)+"""</td>
                                            <td style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);" data-origin-bgcolor="empty" bgcolor="rgb(249, 212, 0)">"""+str(s_leftday)+"""</td>
                                        </tr>
                                        </table>
                                    <Br>
                                    1. 요청일자 :<span style="background-color: rgb(255, 248, 128); color: rgb(0, 0, 0);">"""+str(end_time)+""" 까지</span><Br>
                                    2. 작성예시 <Br>
                                    - 4/3(월) : 1일 사용<Br>
                                    - 5/10(수) : 0.5일 사용<Br>
                                    <Br>
                                    현재 입사일 기준으로 발생한 연(월)차를 발생일 기준 1년 내에 모두 소진해야 하는 것을 원칙으로 하고 있습니다.<Br>
                                    *시행일자 : 22.06.01<Br>
                                    위와 안내드린 바와 같이 사용 기간 내에 연차를 사용하지 않을 경우 자동 소멸 되오니,<Br>
                                    이점 참고하여 잔여 연차가 발생하지 않도록 계획하여 사용해 주시기 바랍니다.<Br>
                                    <Br>
                                    감사합니다.
                                    """
                       
                        
                        driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_body > textarea:nth-child(1)').send_keys(mail_data)
                        
                        driver.find_element(By.CSS_SELECTOR,'#contact_layer_offset_elem > div.editorFrame > div > div.editor_footer.non_resizable > div > button:nth-child(1)').click()
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR,'#sendBtn').click()
                        time.sleep(0.5)
                        driver.find_element(By.CSS_SELECTOR,'#modal-holder > div > div > div > div > button.lw_btn_point').click()
                        time.sleep(5)
        except :
            self.label_msg2.emit("no send") 
        
class QPushButton(QPushButton):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLabel(QLabel):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLineEdit(QLineEdit):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        #layout_2.addWidget(layout_1)

class QComboBox(QComboBox):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class StWidgetForm(QGroupBox):
    """
    위젯 베이스 클래스
    """
    def __init__(self):
        QGroupBox.__init__(self)
        self.box = QBoxLayout(QBoxLayout.TopToBottom)
        self.setLayout(self.box)

class time_ck(QThread) :
    time_msg=pyqtSignal(str)
    label_msg=pyqtSignal(str)
    send_msg=pyqtSignal(str)
    
    def __init__(self, parent) :
        super().__init__(parent)
        self.parent = parent
        
        #self.db_up = Db_update(parent=self)
    
    def run(self) :
        
        today_data = datetime.datetime.now()
        time.sleep(5)
        time_ = time.strftime("%M")
        label_ = today_data.strftime("%Y년 %m월 %d일 ") + time.strftime("%H시 %M분")
        self.label_msg.emit(str(label_))
        while True :
            msg = self.parent.time_d_send_msg
            msg_2 = self.parent.time_d_send_msg_2
            label_ = today_data.strftime("%Y년 %m월 %d일 ") + time.strftime("%H시 %M분")
           
            #schedule.every(1).second.do(self.time_check)
            if msg == "d" and msg_2 =="d":
                if time.strftime("%H:%M:%S") == "11:00:00":
                
                    self.label_msg.emit("DB")
                    self.time_msg.emit("start")
                    time.sleep(1)
                    
                elif  time.strftime("%H:%M:%S") == "12:00:20": 
                    
                    self.label_msg.emit("send")
                    self.send_msg.emit("send")
                    time.sleep(1)
                
                elif time_ != time.strftime("%M"):
                    self.label_msg.emit(str(label_))
                    time.sleep(1)
                    time_ = time.strftime("%M")
                time.sleep(1)
                    
                   
            elif msg != "d" and msg_2 =="d":
            
                db_h_time=self.parent.time_h_insert.currentText()
                db_m_time=self.parent.time_m_insert.currentText()
                
                db_time = db_h_time + ":"+db_m_time+":00"
                
                if time.strftime("%H:%M:%S") == db_time:
                
                    self.label_msg.emit("DB")
                    self.time_msg.emit("start")
                    time.sleep(1)
                elif  time.strftime("%H:%M:%S") == "12:00:20": 
                    
                    self.label_msg.emit("send")
                    self.send_msg.emit("send")
                    time.sleep(1)
                    
                elif time_ != time.strftime("%M"):
                    self.label_msg.emit(str(label_))
                    time.sleep(1)
                    time_ = time.strftime("%M")
                   
                time.sleep(1)
                        
            elif msg == "d" and msg_2 !="d":
                send_h_time=self.parent.time_h_insert2.currentText()
                send_m_time=self.parent.time_m_insert2.currentText()
                send_time = send_h_time + ":"+send_m_time+":00"
                if time.strftime("%H:%M:%S") == send_time:
                #time.strftime("%H:%M:%S") == "12:00:20"
                    self.label_msg.emit("send")
                    self.send_msg.emit("send")
                    time.sleep(1)
                elif time.strftime("%H:%M:%S") == "11:00:00":
                
                    self.label_msg.emit("DB")
                    self.time_msg.emit("start")
                    time.sleep(1)
                
                elif time_ != time.strftime("%M"):
                    self.label_msg.emit(str(label_))
                    time.sleep(1)
                    time_ = time.strftime("%M")
                time.sleep(1)
                
            
            elif msg != "d" and msg_2 !="d":
                send_h_time=self.parent.time_h_insert2.currentText()
                send_m_time=self.parent.time_m_insert2.currentText()
                send_time = send_h_time + ":"+send_m_time+":00"
                db_h_time=self.parent.time_h_insert.currentText()
                db_m_time=self.parent.time_m_insert.currentText()
                
                db_time = db_h_time + ":"+db_m_time+":00"
                
                if time.strftime("%H:%M:%S") == db_time:
                
                    self.label_msg.emit("DB")
                    self.time_msg.emit("start")
                    time.sleep(1)
                    
                elif time.strftime("%H:%M:%S") == send_time :
              
                    self.label_msg.emit("send")
                    self.send_msg.emit("send")
                    time.sleep(1)
                    
                elif time_ != time.strftime("%M"):
                    self.label_msg.emit(str(label_))
                    time.sleep(1)
                    time_ = time.strftime("%M")
                time.sleep(1)
            else :
                time.sleep(1)
                pass

class Main(QDialog,object):
    def __init__(self):
        super().__init__()
        
        self.main()
    
    def main(self):
        
        self.setWindowTitle("연차확인 및 메일발송")
        self.setWindowFlags(Qt.FramelessWindowHint)
       
        screen_rect =QDesktopWidget().availableGeometry()
        self.wi,self.he = screen_rect.width(), screen_rect.height()
        self.setGeometry(self.wi-250,self.he-250,200,200)

        
        self.main_layout = QtWidgets.QHBoxLayout()

        self.stk_w_login = QStackedWidget()

        self.login_stk1 = QWidget()
        self.verticalLayout_1 = QVBoxLayout(self.login_stk1)

        self.login_stk2 = QWidget()
        self.verticalLayout_2 = QVBoxLayout(self.login_stk2)

        self.login_stk3 = QWidget()
        self.verticalLayout_3 = QVBoxLayout(self.login_stk3)

        
        self.id_insert=QLineEdit()
        self.id_label1 = QLabel("ID")
        self.id_label2 = QLabel("@3frame.com")
        self.id_label2.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.id_layout1 = QHBoxLayout()
        self.id_layout2 = QVBoxLayout()
        self.id_layout1.addWidget(self.id_label1,stretch=1)
        self.id_layout1.addWidget(self.id_insert, stretch=3)
        self.id_layout2.addLayout(self.id_layout1)
        self.id_layout2.addWidget(self.id_label2)

        self.pw_insert=QLineEdit()
        self.pw_label1 = QLabel("PW")
        self.pw_layout = QHBoxLayout()     
        self.pw_insert.setEchoMode(QLineEdit.Password)
        self.pw_layout.addWidget(self.pw_label1, stretch=1)
        self.pw_layout.addWidget(self.pw_insert, stretch=3)
        
        self.login_btn = QPushButton("Login")
        self.login_btn.clicked.connect(self.login_ck)
        self.remeber_id = QCheckBox("로그인 기억")
        self.remever_layout=QHBoxLayout()
        txt_file=os.getcwd()+'/ID_PW.txt'
        if os.path.isfile(txt_file):
            self.remeber_id.setCheckState(2)
            self.set_id()
            
        else :
            self.remeber_id.setCheckState(0)
        self.login_btn_layout = QVBoxLayout()
        self.remeber_id2 = QLabel("")
        self.remever_layout.addWidget(self.remeber_id2,stretch=3)
        self.remever_layout.addWidget(self.remeber_id,stretch=1)
        self.login_btn_layout.addLayout(self.remever_layout)
        self.login_btn_layout.addWidget(self.login_btn)
        

        self.verticalLayout_1.addLayout(self.id_layout1,stretch=2)
        self.verticalLayout_1.addLayout(self.id_layout2,stretch=3)
        self.verticalLayout_1.addLayout(self.pw_layout,stretch=2)
        self.verticalLayout_1.addLayout(self.login_btn_layout,stretch=2)
        

        self.stk_w_login.addWidget(self.login_stk1)

        

        self.main_layout.addWidget(self.stk_w_login)
        
        self.setLayout(self.main_layout)
        self.id = self.id_insert.text()
        self.pw = self.pw_insert.text()
        #find = chrome_login(self.id,self.pw)
        #1page 내용

        self.date = QDate.currentDate()
        self.top_layout = QHBoxLayout()
        self.top_btn_layout = QVBoxLayout()
        self.top_btn_layout2 = QVBoxLayout()
        self.ba_set_btn = QPushButton("설정")
        self.ba_btn = QPushButton("ID,PW 재입력")
        self.ba_label = QLabel("")
        self.ba_hide_btn = QPushButton("숨기기")
       # opacity_effect = QGraphicsOpacityEffect(self.ba_btn)
        #opacity_effect.setOpacity(0.5)
        #self.ba_btn.setGraphicsEffect(opacity_effect)
        self.top_btn_layout.addWidget(self.ba_set_btn,stretch=5)
        self.top_btn_layout.addWidget(self.ba_btn,stretch=5)
        
        self.top_btn_layout2.addWidget(self.ba_label,stretch=3)
        self.top_btn_layout2.addWidget(self.ba_hide_btn,stretch=3)
        self.top_layout.addLayout(self.top_btn_layout2,stretch=3)
        self.top_layout.addLayout(self.top_btn_layout,stretch=2)
        self.ba_hide_btn.clicked.connect(self.hide)
        
        
        
        #self.ba_btn.setStyleSheet("background-color: ")
        self.s_btn = QPushButton("검색")

        self.name_layout = QHBoxLayout()
        self.name_label = QLabel("이름")
        self.name_insert=QLineEdit()
        self.name_layout.addWidget(self.name_label,stretch=2)
        self.name_layout.addWidget(self.name_insert,stretch=4)
        self.name_layout.addWidget(self.s_btn,stretch=2)

        self.class_label = QLabel("부서")
        self.class_insert=QtWidgets.QComboBox()
        self.class_layout = QHBoxLayout()
        self.class_layout.addWidget(self.class_label,stretch=2)
        self.class_layout.addWidget(self.class_insert,stretch=7)
        self.class_insert.currentTextChanged.connect(self.combo_search2)

        #self.verticalLayout_2.addLayout(self.top_layout,stretch=1)
        self.verticalLayout_2.addLayout(self.name_layout,stretch=2)
        self.verticalLayout_2.addLayout(self.class_layout,stretch=2)
        
        self.ba_btn.clicked.connect(self.ba_btn_ck)
        self.ba_set_btn.clicked.connect(self.set_btn_ck)
        
        self.stk_w_login.addWidget(self.login_stk2)
        self.s_btn.clicked.connect(self.combo_search)
       


        self.inday_label = QLabel("입사일")
        self.inday_insert=QLabel()
        self.inday_insert.setAlignment(Qt.AlignRight)
        self.inday_layout = QHBoxLayout()
        self.inday_layout.addWidget(self.inday_label,stretch=2)
        self.inday_layout.addWidget(self.inday_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.inday_layout,stretch=2)

        self.rest_label = QLabel("발생연차")
        self.rest_insert=QLabel()
        self.rest_insert.setAlignment(Qt.AlignRight)
        self.rest_layout = QHBoxLayout()
        self.rest_layout.addWidget(self.rest_label,stretch=2)
        self.rest_layout.addWidget(self.rest_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.rest_layout,stretch=2)
        

        self.use_label = QLabel("사용연차")
        self.use_insert=QLabel()
        self.use_insert.setAlignment(Qt.AlignRight)
        self.use_layout = QHBoxLayout()
        self.use_layout.addWidget(self.use_label,stretch=2)
        self.use_layout.addWidget(self.use_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.use_layout,stretch=2)
        
        self.use2_label = QLabel("남은연차")
        self.use2_insert=QLabel()
        self.use2_insert.setAlignment(Qt.AlignRight)
        self.use2_layout = QHBoxLayout()
        self.use2_layout.addWidget(self.use2_label,stretch=2)
        self.use2_layout.addWidget(self.use2_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.use2_layout,stretch=2)
        
        self.send_label = QLabel("메일 발송일")
        self.send_insert=QLabel()
        self.send_insert.setAlignment(Qt.AlignRight)
        self.send_layout = QHBoxLayout()
        self.send_layout.addWidget(self.send_label,stretch=2)
        self.send_layout.addWidget(self.send_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.send_layout,stretch=2)
        self.verticalLayout_2.addLayout(self.top_layout,stretch=1)

        self.db_up = Db_update(parent=self)
        self.db_up.label_msg.connect(self.label_ck)
        self.time_ck =time_ck(parent=self)
        self.time_ck.time_msg.connect(self.time_ck1)
        self.time_ck.label_msg.connect(self.label_ck)
        self.time_ck.send_msg.connect(self.send_ck)
        
        self.time_ck.start()
        self.send_m = send_mail(parent=self)
        self.send_m.label_msg2.connect(self.label_ck)

        self.stk_w_login.addWidget(self.login_stk3)
        self.time_d_label = QLabel("DB업데이트 시간 설정")
        self.time2_d_label = QLabel("시")
        self.time2_d_label2 = QLabel("분")
        self.time_d_ck = QCheckBox("기본값")
        self.time_d_ck.setCheckState(2)
        self.time_d_ck.stateChanged.connect(self.time_d_ck_)
        self.time_d_send_msg="d"
        self.time_d_layout = QHBoxLayout()
        self.time_d_layout.addWidget(self.time_d_label,stretch=2)
        
        self.time_d_layout.addWidget(self.time_d_ck,stretch=2)
        
        self.time_h_insert=QtWidgets.QComboBox()
        self.time_m_insert=QtWidgets.QComboBox()
        self.time_set_layout = QHBoxLayout()
        self.time_set_layout.addWidget(self.time_h_insert,stretch=2)
        self.time_set_layout.addWidget(self.time2_d_label,stretch=2)
        self.time_set_layout.addWidget(self.time_m_insert,stretch=2)
        self.time_set_layout.addWidget(self.time2_d_label2,stretch=2)
        self.time_layout = QVBoxLayout()
        self.time_layout.addLayout(self.time_d_layout,stretch=2)
        self.time_layout.addLayout(self.time_set_layout,stretch=2)
        self.verticalLayout_3.addLayout(self.time_layout,stretch=2)
        
        
        self.stk_w_login.addWidget(self.login_stk3)
        self.time_d_label2 = QLabel("메일발송 시간 설정")
        
        self.time_d_ck2 = QCheckBox("기본값")
        self.time_d_ck2.setCheckState(2)
        self.time_d_ck2.stateChanged.connect(self.time_d_ck_2)
        self.send_time2_d_label = QLabel("시")
        self.send_time2_d_label2 = QLabel("분")
        self.time_d_send_msg_2="d"
        self.time_d_layout2 = QHBoxLayout()
        
        
        self.time_d_layout2.addWidget(self.time_d_label2,stretch=2)
        self.time_d_layout2.addWidget(self.time_d_ck2,stretch=2)
        self.time_h_insert2=QtWidgets.QComboBox()
        self.time_m_insert2=QtWidgets.QComboBox()
        self.time_set_layout2 = QHBoxLayout()
        self.time_set_layout2.addWidget(self.time_h_insert2,stretch=2)
        self.time_set_layout2.addWidget(self.send_time2_d_label,stretch=2)
        self.time_set_layout2.addWidget(self.time_m_insert2,stretch=2)
        self.time_set_layout2.addWidget(self.send_time2_d_label2,stretch=2)
        self.time_layout2 = QVBoxLayout()
        self.time_layout2.addLayout(self.time_d_layout2,stretch=2)
        self.time_layout2.addLayout(self.time_set_layout2,stretch=2)
        self.verticalLayout_3.addLayout(self.time_layout2,stretch=2)
        
        self.mid_layout = QHBoxLayout()
        self.verticalLayout_3.addLayout(self.mid_layout,stretch=3)
        
        self.top_layout2 = QHBoxLayout()
        self.top_btn_layout2 = QVBoxLayout()
        self.ba_set_btn2 = QPushButton("검색")
        self.ba_btn2 = QPushButton("ID,PW 재입력")
        
        self.ba_set_btn2.clicked.connect(self.mi_btn_ck)
        self.ba_btn2.clicked.connect(self.ba_btn_ck)
        self.ba_label2 = QLabel("")
        self.top_btn_layout2.addWidget(self.ba_set_btn2,stretch=5)
        self.top_btn_layout2.addWidget(self.ba_btn2,stretch=5)
        
        self.top_layout2.addWidget(self.ba_label2,stretch=3)
        self.top_layout2.addLayout(self.top_btn_layout2,stretch=2)
        self.verticalLayout_3.addLayout(self.top_layout2,stretch=3)
        
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        
        show_action = QAction("Show", self)
        quit_action = QAction("Exit", self)
        hide_action = QAction("Hide", self)
        
        show_action.triggered.connect(self.showing)
        
        hide_action.triggered.connect(self.hide)
        quit_action.triggered.connect(self.exit)
        tray_menu = QMenu()
        tray_menu.addAction(show_action)
        tray_menu.addAction(hide_action)
        tray_menu.addAction(quit_action)
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()
    
    def mousePressEvent(self, event):
        try :
            if event.button() == Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)
        except :
            pass
    
    def mouseMoveEvent(self, event):
        try :
            if self.offset is not None and event.buttons() == Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)
        except :
            pass

    def mouseReleaseEvent(self, event):
        try :
            self.offset = None
            super().mouseReleaseEvent(event)
        except :
            pass   
    def set_id(self):
        id_pw=open(os.getcwd()+'/ID_PW.txt',"r",encoding="utf-8")
        ID_d =id_pw.readlines()
        self.id_insert.setText(ID_d[0].rstrip('\n'))
        print(ID_d[0])
        self.pw_insert.setText(ID_d[1].rstrip('\n'))
        print(ID_d[1])
        
    def save_id(self):
        txt_file=os.getcwd()+'/ID_PW.txt'
        
        if os.path.isfile(txt_file):
            id_pw=open(os.getcwd()+'/ID_PW.txt',"r",encoding="utf-8")
            r_id =id_pw.readline().rstrip('\n')
            if r_id ==self.id_insert.text():
                ID_d =id_pw.readlines()
                self.id_insert.setText(ID_d[0].rstrip('\n'))
                self.pw_insert.setText(ID_d[1].rstrip('\n'))
            if id_pw.readline() !=self.id_insert.text():
                outFp = open(os.getcwd()+'/ID_PW.txt',"w",encoding="utf-8")
                src_id = self.id_insert.text()
               # s_id =bz2.compress(str(src_id))
                #print(s_id)
                src_pw = self.id_insert.text()
                #outFp.writelines(bz2.compress(str(src_id)))
              # outFp.writelines(bz2.compress(str(src_pw)))
            id_pw.close()
        else :
            outFp = open(os.getcwd()+'/ID_PW.txt',"w",encoding="utf-8")
            src_id = self.id_insert.text()
            src_pw = self.id_insert.text()
            
            
           # outFp.writelines(bz2.compress(str(src_pw)))
            
            outFp.close()
    def hider(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().bottomRight()
        qr.moveBottomLeft(cp)
        self.move(qr.bottomRight())   
    
    def showing(self) :
        self.setGeometry(self.wi-300,self.he-300,300,300)
        self.show()

    def exit(self) :
        pid = os.getpid()
        os.kill(pid, 2)    

    @pyqtSlot(str)
    def send_ck(self,msg) :
        
        if msg == "send":
            
            self.send_m.start()

    @pyqtSlot(str)
    def time_ck1(self,msg) :
        
        if msg == "start":
            
            self.db_up.start()
    
    @pyqtSlot(str)
    def label_ck(self,msg) :
        
        if msg == "DB":
            
            self.ba_label.setText("DB 업데이트 중")  

        elif msg == "DB up":
            self.ba_label.setText("DB 업데이트 완료")  

        elif msg =="send":
           
            self.ba_label.setText("메일 발송 중")
        
        elif msg == "no send" :
            self.ba_label.setText("메일 발송 건 없음")
            
        elif msg == "com send" :
            self.ba_label.setText("메일 발송 완료")
        else :

            self.ba_label.setText(msg)
    

        
    def login_ck(self):
        if self.id_insert.text() == "" and self.pw_insert.text() == "":
            QMessageBox.information(self,"확인","ID,PW를 입력해주세요")
        elif self.id_insert.text() == "" :
            QMessageBox.information(self,"확인","ID를 입력해주세요")
        elif self.pw_insert.text() == "" :
            QMessageBox.information(self,"확인","PW를 입력해주세요")
        else :
            self.setGeometry(self.wi-300,self.he-300,300,300)
            self.stk_w_login.setCurrentWidget(self.login_stk2)
        self.save_id()
    def ba_btn_ck(self) :
        self.stk_w_login.setCurrentWidget(self.login_stk1)
    def mi_btn_ck(self) :
        self.stk_w_login.setCurrentWidget(self.login_stk2)
    def set_btn_ck(self) :
        self.stk_w_login.setCurrentWidget(self.login_stk3)
    
    

    def combo_search2(self,item):
        try :
            name = self.name_insert.text()
            if name == "" :
                #QMessageBox.information(self,"확인","이름을 입력해주세요")
                self.clear_value()
            else :
                
                for row in database_ws.iter_rows(min_row=2):
                    if row[1].value == name :
                        if row[3].value == item:
                            #in_d = row[4].value
                            self.inday_insert.setText(str(row[4].value))
                            #r_d = row[5].value
                            self.rest_insert.setText(str(row[5].value))
                            #u_d=row[6].value
                            self.use_insert.setText(str(row[6].value)+"일")
                            #l_d=row[7].value
                            self.use2_insert.setText(str(row[7].value)+"일")
                #self.inday_insert.setText(str(in_d))
                #self.rest_insert.setText(str(r_d))
                #self.use_insert.setText(str(u_d)+"일")
                #self.use2_insert.setText(str(l_d)+"일")
            
                name = self.name_insert.text()
                for row1 in database_ws2.iter_rows(min_row=2):
                    if row1[1].value == name :
                        if row1[3].value == item:
                            #s_d = row1[4].value
                            self.send_insert.setText(str(row1[4].value))
                #self.send_insert.setText(str(s_d))
        except OSError:
            pass


    def combo_search(self):
        try :
            name = self.name_insert.text()
            self.clearComboBoxItem()
            self.clear_value()
            time.sleep(0.2)
            if name =="" :
                QMessageBox.information(self,"확인","이름을 입력해주세요")
                self.clear_value()
            else :    
                for row in database_ws.iter_rows(min_row=2):
                        
                    if row[1].value == name :
                    
                        self.h_class =row[3].value 
                        self.addComboBoxItem()
                
                    else :
                        
                        pass  
        except OSError :
            pass
    
    def clear_value(self) :
        self.inday_insert.setText("")
        self.rest_insert.setText("")
        self.use_insert.setText("")
        self.use2_insert.setText("")
        self.send_insert.setText("")
    def clearComboBoxItem(self) :
        self.class_insert.clear()
    
    def addComboBoxItem(self) :
        self.class_insert.addItem(self.h_class)
    
    def time_d_ck_(self,state):
        
        h_time = ['00','01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23']
        m_time = ['00','05','10','15','20','25','30','35','40','45','50','55']
        if state == 2:
            self.time_h_insert.clear()
            self.time_m_insert.clear()
                   
        else :
            for i in h_time :
                
                self.time_h_insert.addItem(str(i))
                
            for l in m_time :
                self.time_m_insert.addItem(str(l))
            
            self.time_d_send_msg = ""

    def time_d_ck_2(self,state):
        
        h_time = ['00','01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23']
        m_time = ['00','05','10','15','20','25','30','35','40','45','50','55']
        if state == 2:
            self.time_h_insert2.clear()
            self.time_m_insert2.clear()
            
        else :
            for i in h_time :
                
                self.time_h_insert2.addItem(str(i))
                
            for l in m_time :
                self.time_m_insert2.addItem(str(l))
            
            self.time_d_send_msg_2 = ""
    
    def back_id (self) :
        inFileName = "ID_PW_bak.txt"
        outFileName = "ID_PW_re.txt "
        secuYN = "1. 암호화 2.복호화: "
        secu = 0
        if secuYN == '1':
            secu = 100
        elif secuYN == '2':
            secu = -100
        else:
            secu = 0
            
        inFp = open(inFileName, "r", encoding='utf-8')
        outFp = open(outFileName, "w", encoding="utf-8")
        for inStr in inFp.readlines():
            out_str = ""
            for s in inStr:
                tmp = ord(s) + secu
                out_str += chr(tmp)
            outFp.writelines(out_str)

        inFp.close()
        outFp.close()
            
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = Main()
    main.show()
    sys.exit(app.exec_())
