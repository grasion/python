import sys
from types import NoneType
from xmlrpc.client import DateTime
from PyQt5.QtWidgets import *
import os
from PyQt5.QtCore import *
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtGui import QStandardItem
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QIcon
from openpyxl import load_workbook
import datetime
from dateutil.relativedelta import relativedelta

database_wb=load_workbook(filename=os.getcwd()+'/사원명부.xlsx',data_only=True)
database_ws=database_wb['명부']
database_ws_end=database_wb['퇴사자']
today_data = datetime.datetime.now()


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))    
    return os.path.join(base_path, relative_path)

### 사이즈 정책을 설정한 새로운 class를 생성합니다. ###
class QPushButton(QPushButton):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLabel(QLabel):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLineEdit(QLineEdit):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        #layout_2.addWidget(layout_1)

class QComboBox(QComboBox):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class StWidgetForm(QGroupBox):
    """
    위젯 베이스 클래스
    """
    def __init__(self):
        QGroupBox.__init__(self)
        self.box = QBoxLayout(QBoxLayout.TopToBottom)
        self.setLayout(self.box)
        
class search(StWidgetForm,QDialog):
    """
    검색 그룹
    """
    def __init__(self):
        
        #layout_equation_solution = QFormLayout()
        super(search, self).__init__()
        self.setTitle("검색")
        
        #self.box =  QtWidgets.QWidget()
        #self.setGeometry(0, 0, 700, 700)
        #self.box = QHBoxLayout(sub_w)
        self.main_layout = QVBoxLayout(self)
        self.stack_layout2 = QtWidgets.QStackedWidget(self)
        self.stk_w2 = QStackedWidget()
        
        
        layout_1=QHBoxLayout()
        layout_2=QHBoxLayout()
        layout_3=QHBoxLayout()
        layout_4=QHBoxLayout()
        layout_5=QHBoxLayout()
        layout_6=QHBoxLayout()
        layout_7=QVBoxLayout()
        layout_8=QHBoxLayout()
        layout_9=QHBoxLayout()
        layout_10=QHBoxLayout()
        layout_11=QHBoxLayout()
        layout_12=QHBoxLayout()
        layout_13=QHBoxLayout()
        layout_14=QHBoxLayout()
        
        self.comboBox = QtWidgets.QComboBox()
        self.name_search = QLabel("이름 : ")
        job_search = QLabel("부서명 : ")
        
        in_search = QLabel("입사일 : ")
        b_search = QLabel("생   일 : ")
        etc_search = QLabel("기타 중요일 : ")
        etc_con_search = QLabel("중요일 내용 : ")
        self.button_1 = QPushButton("검색")
        mon_search = QLabel("달   수 : ")
        self.mon_in = QLabel(" ")
        day_search = QLabel("일   수 : ")
        self.day_in = QLabel(" ")
        self.etclabel = QLabel(" 부서 수정시 체크박스 체크 ")
        thmon_search = QLabel("3 개 월 : ")
        self.thmon_in = QLabel(" ")
        self.button_2 = QPushButton("수정")
        self.button_3 = QPushButton("퇴사")        
        #button_2 = QPushButton("수정")
        self.name_equation = QLineEdit("")
        self.in_equation = QLineEdit("")
        self.b_equation = QLineEdit("")
        self.etc_con_equation = QLineEdit("")
        self.etc_equation = QLineEdit("")
        checkbox = QCheckBox("수정")
        self.page_1 = QWidget()
        
        self.verticalLayout_1 = QHBoxLayout(self.page_1)
        self.page_2 = QWidget()
        
        self.verticalLayout_2 = QHBoxLayout(self.page_2)
        #self.stk_w2.addWidget(QWidget())
        #self.stk_w2.addWidget(QWidget())
       
        job_search2 = QLabel("신규 부서명 : ")
        self.job_equation = QLineEdit(" ")
        self.verticalLayout_1.addWidget(self.etclabel)
        self.verticalLayout_2.addWidget(job_search2, stretch=1)
        self.verticalLayout_2.addWidget(self.job_equation, stretch=5)
        #self.page_2.addWidget(layout_14)
        self.stk_w2.addWidget(self.page_1)
        self.stk_w2.addWidget(self.page_2)
        layout_13.addWidget(self.stk_w2)
        checkbox.stateChanged.connect(self.checkbox_changed)
        #self.stack_layout2.addWidget(self.stk_w2)

        #print(self.in_equation.sizeHint())
        #print(self.in_equation.sizePolicy().horizontalPolicy())
        layout_1.addWidget(self.name_search, stretch=1)
        layout_1.addWidget(self.name_equation, stretch=9)
        layout_1.addWidget(self.button_1, stretch=1)
        layout_2.addWidget(job_search, stretch=1)
        layout_2.addWidget(self.comboBox, stretch=7)
        layout_2.addWidget(checkbox, stretch=1)
        layout_3.addWidget(in_search, stretch=1)
        layout_3.addWidget(self.in_equation, stretch=9)
        layout_4.addWidget(b_search, stretch=1)
        layout_4.addWidget(self.b_equation, stretch=9)
        layout_5.addWidget(etc_con_search, stretch=1)
        layout_5.addWidget(self.etc_con_equation, stretch=9)
        layout_6.addWidget(etc_search, stretch=1)
        layout_6.addWidget(self.etc_equation, stretch=9)
        layout_8.addWidget(mon_search)
        layout_8.addWidget(self.mon_in)
        layout_9.addWidget(day_search)
        layout_9.addWidget(self.day_in)
        layout_10.addWidget(thmon_search)
        layout_10.addWidget(self.thmon_in)
        layout_7.addLayout(layout_10)
        layout_7.addLayout(layout_8)
        layout_7.addLayout(layout_9)
        layout_11.addWidget(self.button_2, stretch=1)
        layout_11.addWidget(self.button_3, stretch=1)
        layout_12.addWidget(self.stk_w2, stretch=1)
        #layout_1.addLayout(btn_layout_1)
        self.box.addLayout(layout_1)
        self.box.addLayout(layout_2)
        self.box.addLayout(layout_13)
        #self.box.addLayout(layout_12)
        self.box.addLayout(layout_3)
        self.box.addLayout(layout_7)
        self.box.addLayout(layout_4)
        self.box.addLayout(layout_5)
        self.box.addLayout(layout_6)
        self.box.addLayout(layout_11)
        self.button_1.clicked.connect(self.inday_serch)
        self.button_2.clicked.connect(self.revise_con)
        self.button_3.clicked.connect(self.del_h)
        self.comboBox.currentTextChanged.connect(self.combo_search)
        #layout_equation_solution.addRow(name_search, self.name_equation)
        #layout_equation_solution.addRow(button_1,button_2)
        #layout_equation_solution.addRow(in_search, self.in_equation)
        #self.box.addLayout(layout_1)
    def clear_line(self) :   
        self.etc_con_equation.clear()
        self.name_equation.clear()
        self.in_equation.clear()
        self.b_equation.clear()
        self.etc_con_equation.clear()
        self.etc_equation.clear()
        self.thmon_in.clear()
        self.mon_in.clear()
        self.day_in.clear()
        self.job_equation.clear()
        self.clearComboBoxItem()

    def del_h(self) :
        if self.name_equation.text() == '' :
            self.clear_line()
            QMessageBox.information(self,"확인","검색을 먼저 해주시기 바랍니다.")
        else :
            name = self.name_equation.text()
            item = self.comboBox.currentText()
            self.list_num_f = 1
            row_max = database_ws_end.max_row + 1 
            for row in database_ws.iter_rows(min_row=2):
                if row[0].value == name :
                    if row[2].value == item :
                        con_row = row
                        self.list_num_f = self.list_num_f+1
                    else :
                        self.list_num_f = self.list_num_f+1
                        pass
                else :
                    self.list_num_f = self.list_num_f+1
                    pass
            for h in range(0,len(con_row)) :
                database_ws_end.cell(row=row_max,column=int(h)+1).value =con_row[h].value
        #row_max.value=con_row
            database_ws.delete_rows(self.list_num_f)
            database_wb.save(os.getcwd()+'/사원명부.xlsx')
            self.job_equation.setText("")
            self.clear_line()
            QMessageBox.information(self,"확인","퇴사처리 되었습니다.")
    

    def checkbox_changed(self, state):
        if state == 0 :
            self.stk_w2.setCurrentIndex(0)
        else :
            self.stk_w2.setCurrentIndex(1)

    def revise_con(self):

        if self.name_equation.text() == '' :
            self.clear_line()
            QMessageBox.information(self,"확인","검색을 먼저 해주시기 바랍니다.")
        else :
            name = self.name_equation.text()
       
            item = self.comboBox.currentText()
                
            for row in database_ws.iter_rows(min_row=2):
                
                if row[0].value == name :

                    if row[2].value == item :
                    
                        if type(row[1].value) is NoneType :
                            in_day = row[1].value
                        else :
                            in_day = row[1].value.strftime("%Y년 %m월 %d일")
                    
                        if type(row[3].value) is NoneType :   
                            b_day=row[3].value
                        else :
                            b_day=row[3].value.strftime("%Y년 %m월 %d일")
                        
                        etc_con=row[5].value

                        if type(row[4].value) is NoneType :
                            etc_day=row[4].value
                        else :
                            etc_day=row[4].value.strftime("%Y년 %m월 %d일")

                        self.revise_in_day = self.in_equation.text()
                        self.revise_b_day = self.b_equation.text()
                        self.revise_etc_day = self.etc_equation.text()
                        self.revise_etc_con = self.etc_con_equation.text()

                        if in_day==self.in_equation.text() or self.in_equation.text() == 'None':
                            pass
                        else :
                            self.revise_in_day = datetime.datetime.strptime(self.revise_in_day,'%Y%m%d')
                            row[1].value = self.revise_in_day

                        if b_day == self.b_equation.text() or self.b_equation.text() == 'None':
                            pass
                        else :
                            self.revise_b_day = datetime.datetime.strptime(self.revise_b_day,'%Y%m%d')
                            row[3].value = self.revise_b_day
                    
                        if etc_con ==self.etc_con_equation.text() or self.etc_con_equation.text() == 'None':
                            pass
                        else :
                            row[5].value = self.revise_etc_con 
                    
                    
                        if etc_day == self.etc_equation.text() or self.etc_equation.text() == 'None':
                            pass
                        
                        else :    
                            self.revise_etc_day = datetime.datetime.strptime(self.revise_etc_day,'%Y%m%d')
                            row[4].value = self.revise_etc_day

                        if self.job_equation.text() == " "or self.job_equation.text() == 'None':
                            pass
                        else :
                            row[2].value = self.job_equation.text()
                        
                    else :
                        pass
                else :
                
                    pass  
        
        
            database_wb.save(os.getcwd()+'/사원명부.xlsx')
            self.clear_line()
            QMessageBox.information(self,"확인","수정되었습니다..")

    def inday_serch(self):
        
        name = self.name_equation.text()
        self.clearComboBoxItem()
        
        for row in database_ws.iter_rows(min_row=2):
                
            if row[0].value == name :
               
                self.h_class =row[2].value 
                self.addComboBoxItem()
           
            else :
                
                pass  
    def combo_search(self,item):
        
        name = self.name_equation.text()
       # self.clearComboBoxItem()
       
        for row in database_ws.iter_rows(min_row=2):
            
            if row[0].value == name :

                if row[2].value == item :
                    
                    #etc_day=row[4].value
                    self.h_class =row[2].value 
                    
                    if type(row[1].value) is NoneType :
                        self.in_day = row[1].value
                        m_total_day = '입사일을 입력해주세요'
                        total_day = '입사일을 입력해주세요'
                        after_th_month = '입사일을 입력해주세요'
                    else :
                        self.in_day = row[1].value
                        total_day=str((today_data-self.in_day).days)+'일'
                        m_total_day =str((today_data.year - self.in_day.year) * 12 + today_data.month - self.in_day.month )+'개월'
                        after_th_month = (self.in_day + relativedelta(months=3)).strftime("%Y년 %m월 %d일")
                        self.in_day = row[1].value.strftime("%Y년 %m월 %d일")
                    if type(row[3].value) is NoneType :
                        self.b_day = row[3].value
                        
                    else :
                        self.b_day = row[3].value.strftime("%Y년 %m월 %d일")

                    
                    self.etc_con=row[5].value
                    
                    if type(row[4].value) is NoneType :
                        self.etc_day = row[4].value
                        self.etc_equation.setText(str(self.etc_day))

                    if type(row[4].value) is datetime.datetime :
                        self.etc_day = row[4].value
                    
                        self.etc_equation.setText(str(self.etc_day.strftime("%Y년 %m월 %d일")))

                    h_class =row[2].value 
                    self.in_equation.setText(str(self.in_day))
                    self.b_equation.setText(str(self.b_day))
                    self.mon_in.setText(str(m_total_day))
                    self.day_in.setText(str(total_day))
                    self.thmon_in.setText(str(after_th_month))
                    self.etc_con_equation.setText(str(self.etc_con))
                else :
                    
                    pass
            else :
                
                pass  

    def clearComboBoxItem(self) :
        self.comboBox.clear()
    
    def addComboBoxItem(self) :
        self.comboBox.addItem(self.h_class)
        
class search_in(StWidgetForm,QDialog):
    """
    입력 그룹
    """
    def __init__(self):
        #layout_equation_solution = QFormLayout()
        super(search_in, self).__init__()
    
        self.setTitle("입력")
        layout_1=QHBoxLayout()
        layout_2=QHBoxLayout()
        layout_3=QHBoxLayout()
        layout_4=QHBoxLayout()
        layout_5=QHBoxLayout()
        layout_6=QHBoxLayout()
        layout_7=QVBoxLayout()
        layout_8=QHBoxLayout()
        layout_9=QHBoxLayout()
        layout_10=QHBoxLayout()
        layout_11=QHBoxLayout()
        self.jobbox = QLineEdit("")
        self.name_search = QLabel("이   름 : ")
        job_search = QLabel("부서명 : ")
        in_search = QLabel("입사일 : ")
        b_search = QLabel("생   일 : ")
        etc_search = QLabel("기타 중요일 : ")
        etc_con_search = QLabel("중요일 내용 : ")
        self.button_1 = QPushButton("입력")
        mon_search = QLabel("달수 : ")
        self.mon_in = QLabel(" ")
        day_search = QLabel("일수 : ")
        self.day_in = QLabel(" ")
        thmon_search = QLabel("3개월 : ")
        self.thmon_in = QLabel(" ")            
        self.name_equation = QLineEdit("")
        self.in_equation = QLineEdit("")
        self.b_equation = QLineEdit("")
        self.etc_con_equation = QLineEdit("")
        self.etc_equation = QLineEdit("")
        

        layout_1.addWidget(self.name_search, stretch=1)
        layout_1.addWidget(self.name_equation, stretch=9)
        layout_11.addWidget(self.button_1, stretch=9)
        layout_2.addWidget(job_search, stretch=1)
        layout_2.addWidget(self.jobbox, stretch=7)
        layout_3.addWidget(in_search, stretch=1)
        layout_3.addWidget(self.in_equation, stretch=9)
        layout_4.addWidget(b_search, stretch=1)
        layout_4.addWidget(self.b_equation, stretch=9)
        layout_5.addWidget(etc_con_search, stretch=1)
        layout_5.addWidget(self.etc_con_equation, stretch=9)
        layout_6.addWidget(etc_search, stretch=1)
        layout_6.addWidget(self.etc_equation, stretch=9)
        layout_8.addWidget(mon_search)
        layout_8.addWidget(self.mon_in)
        layout_9.addWidget(day_search)
        layout_9.addWidget(self.day_in)
        layout_10.addWidget(thmon_search)
        layout_10.addWidget(self.thmon_in)
        

        self.box.addLayout(layout_1)
        self.box.addLayout(layout_2)
        self.box.addLayout(layout_3)
        self.box.addLayout(layout_7)
        self.box.addLayout(layout_4)
        self.box.addLayout(layout_5)
        self.box.addLayout(layout_6)
        self.box.addLayout(layout_11)
        self.button_1.clicked.connect(self.insert_h)
       
    def clear_line(self) :   
        
        self.name_equation.clear()
        self.in_equation.clear()
        self.b_equation.clear()
        self.etc_con_equation.clear()
        self.etc_equation.clear()
        
        self.jobbox.clear()

    def insert_h(self) :
        name = self.name_equation.text()
        in_day = self.in_equation.text()
        h_class = self.jobbox.text()
        b_day = self.b_equation.text()
        etc_con=self.etc_con_equation.text()
        etc_day = self.etc_equation.text()
        
        if in_day == '':
            in_day=None
        else :
            in_day=datetime.datetime.strptime(in_day,'%Y%m%d')

        if b_day == '':
            b_day=None
        else :
            b_day=datetime.datetime.strptime(b_day,'%Y%m%d')

        if etc_day == '':
            etc_day=None
        else :
            etc_day=datetime.datetime.strptime(etc_day,'%Y%m%d')

        row_max = database_ws.max_row + 1 
        database_ws.cell(row=row_max,column=1).value = name
        database_ws.cell(row=row_max,column=2).value = in_day
        database_ws.cell(row=row_max,column=3).value = h_class
        database_ws.cell(row=row_max,column=4).value = b_day
        database_ws.cell(row=row_max,column=5).value = etc_day
        database_ws.cell(row=row_max,column=6).value = etc_con
        database_wb.save(os.getcwd()+'/사원명부.xlsx')
        self.clear_line()
        QMessageBox.information(self,"확인","입력되었습니다.")

class base(StWidgetForm):
    
    def __init__(self):
        layout_equation_solution = QFormLayout()
        super(base, self).__init__()
        self.setTitle("메뉴")
        #self.box1=QBoxLayout(QBoxLayout.LeftToRight)
      

class Main(QDialog,object):
    def __init__(self):
        #super().__init__()
        #self.main()
        QWidget.__init__(self, flags=Qt.Widget)
        self.stk_w = QStackedWidget(self)
        
        self.setWindowTitle("3Frame")
        self.setWindowIcon(QIcon('icon.png'))
        main_w =  QtWidgets.QWidget()
        #self.setGeometry(0, 0, 700, 700)
        self.widget_layout = QVBoxLayout(main_w)
        #main_layout = QBoxLayout(QBoxLayout.TopToBottom)
        self.stack_layout = QtWidgets.QStackedWidget(main_w)
        self.main_layout = QtWidgets.QHBoxLayout(main_w)
        #main_layout = QVBoxLayout()
        
        
        

        equation = QLineEdit("")
        solution = QLineEdit("")
        name_search = QLabel("이름 : ")
        print(equation.sizeHint())
        print(equation.sizePolicy().horizontalPolicy())

        
        self.stk_w.addWidget(search())
        self.stk_w.addWidget(search_in())
        self.stack_layout.addWidget(self.stk_w)

        ### 사칙연산 버튼을 layout_operation 레이아웃에 추가
        group = QGroupBox()
        box = QBoxLayout(QBoxLayout.LeftToRight)
        group.setLayout(box)
        group.setTitle("메뉴")
        self.main_layout.addWidget(group)

        button_1 = QPushButton("검색")
        button_2 = QPushButton("입력")
        #button_3 = QPushButton("입력")
        box.addWidget(button_1)
        box.addWidget(button_2)
        #box.addWidget(button_3)
       
        self.widget_layout.addWidget(self.stack_layout, stretch=6)
        self.widget_layout.addLayout(self.main_layout, stretch=1)
        button_1.clicked.connect(self.fnc_btn_1)
        button_2.clicked.connect(self.fnc_btn_2)

        self.setLayout(self.widget_layout)
        self.move(0,0)
       # self.show()

    def fnc_btn_1(self):
        self.stk_w.setCurrentIndex(0)                       # 현재 Index(페이지)를 1로
    def fnc_btn_2(self):
        self.stk_w.setCurrentIndex(1)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = Main()
    main.show()
    sys.exit(app.exec_())

