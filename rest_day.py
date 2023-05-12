import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtWidgets
import os
import time
import sys
from openpyxl import load_workbook
import schedule
import xlwings as xw
import pandas as pd
import pythoncom
import datetime

database_wb=load_workbook(filename=os.getcwd()+'/사원명부.xlsm',data_only=True)
database_ws=database_wb['명단']
database_ws2=database_wb['발송여부']

class chrome_update():
    def ch_update():
        chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
        driver_path = f'./{chrome_ver}/chromedriver.exe'
        if os.path.exists(driver_path):
            print(f"chrom driver is insatlled: {driver_path}")
        else:
            print(f"install the chrome driver(ver: {chrome_ver})")
            chromedriver_autoinstaller.install(True)
        driver=webdriver.Chrome()
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        options.add_experimental_option("prefs",{
            "download.default_directory": os.getcwd(),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
        })
        #options.add_argument("headless")

class Db_update(QThread):
    label_msg=pyqtSignal(str)
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent
       
    def run(self):
        l_id = self.parent.id_insert.text()+"@3frame.com"
        l_pw = self.parent.pw_insert.text()
        chrome_update.ch_update
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches',['enable-logging'])
        options.add_experimental_option("prefs",{
            "download.default_directory": os.getcwd(),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
            })
        options.add_argument("headless")
        driver = webdriver.Chrome(options=options)
        driver.get("https://3frame.ncpworkplace.com/")


        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'user')))
        elem_id = driver.find_element(By.ID, 'user')
        elem_id.send_keys(l_id)

        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'password')))
        elem_pw = driver.find_element(By.ID, 'password')
        elem_pw.send_keys(l_pw)

        login_btn = driver.find_element(By.XPATH, '//*[@id="loginBtn"]')
        login_btn.click()


        time.sleep(5)

        
        self.rest_day(driver)
        
        time.sleep(2)
        pythoncom.CoInitialize()
        app = xw.App(visible=False)
        wb2 = xw.Book(os.getcwd()+'/사원명부.xlsm')
        macro = wb2.macro('filter')
        macro()
        time.sleep(2)
        app.kill()
        self.label_msg.emit("DB up")
                        
    def rest_day(self,driver) :
        time.sleep(3)
        driver.get("https://3frame.ncpworkplace.com/hrs/manager/annualLeave/list")
        time.sleep(3)
        lis_btn = driver.find_element(By.CSS_SELECTOR,'#tap_area1 > div.form-row.row > button')
        lis_btn.click()

class send_mail(QThread):
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent

    def run(self) :

        
        time.sleep(2)

        wb3=load_workbook(filename=os.getcwd()+'/'+time.strftime("%Y%m%d")+'/발송용.xlsx',data_only=True)
        ws3_1 = wb3[wb3.sheetnames[0]]
        ws3_2 = wb3[wb3.sheetnames[1]]
        send_list = len(pd.DataFrame(ws3_1))
        re_send_list = len(pd.DataFrame(ws3_2))

        if send_list >=2 :
            for 행 in ws3_1.iter_rows(min_row=2):
                s_mail = 행[2].value
                s_name = 행[1].value
                s_class = 행[3].value
                s_restday = 행[5].value
                s_useday = 행[6].value
                s_leftday=행[7].value
                df = pd.DataFrame({'부서':[s_class],
                                   '이름':[s_name],
                                   '발생연차':[s_restday],
                                   '사용연차':[s_useday],
                                   '미사용연차':[s_leftday]})
        if re_send_list >=2 :
            for 행 in ws3_2.iter_rows(min_row=2):
                s_mail = 행[2].value
                s_name = 행[1].value
                s_class = 행[3].value
                s_restday = 행[5].value
                s_useday = 행[6].value
                s_leftday=행[7].value
                df = pd.DataFrame({'부서':[s_class],
                                   '이름':[s_name],
                                   '발생연차':[s_restday],
                                   '사용연차':[s_useday],
                                   '미사용연차':[s_leftday]})
                
                print(df)
                
        

class QPushButton(QPushButton):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLabel(QLabel):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class QLineEdit(QLineEdit):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        #layout_2.addWidget(layout_1)

class QComboBox(QComboBox):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

class StWidgetForm(QGroupBox):
    """
    위젯 베이스 클래스
    """
    def __init__(self):
        QGroupBox.__init__(self)
        self.box = QBoxLayout(QBoxLayout.TopToBottom)
        self.setLayout(self.box)

class time_ck(QThread) :
    time_msg=pyqtSignal(str)
    label_msg=pyqtSignal(str)
    send_msg=pyqtSignal(str)
    def __init__(self, parent) :
        super().__init__(parent)
        self.parent = parent
        #self.db_up = Db_update(parent=self)
    def run(self) :
        today_data = datetime.datetime.now()
        time_ = time.strftime("%M")
        label_ = today_data.strftime("%Y년 %m월 %d일 ") + time.strftime("%H시 %M분")
        self.label_msg.emit(str(label_))
        while True :
            label_ = today_data.strftime("%Y년 %m월 %d일 ") + time.strftime("%H시 %M분")
            print(time.strftime("%H:%M:%S"))
            #schedule.every(1).second.do(self.time_check)
            if time.strftime("%H:%M:%S") == "11:46:20":
                
                self.label_msg.emit("DB")
                self.time_msg.emit("start")
                #self.db_up.start()
                time.sleep(1)
            if time.strftime("%H:%M:%S") == "10:45:20":
                self.label_msg.emit("send")
                self.send_msg.emit("send")
                time.sleep(1)
            if time_ != time.strftime("%M"):
                self.label_msg.emit(str(label_))
                time.sleep(1)
            else :
                time.sleep(1)
                pass

class Main(QDialog,object):
    def __init__(self):
        super().__init__()
        
        self.main()

    
    
    def main(self):
        
        self.setWindowTitle("연차확인 및 메일발송")
        
        self.main_layout = QtWidgets.QHBoxLayout()

        self.stk_w_login = QStackedWidget()

        self.login_stk1 = QWidget()
        self.verticalLayout_1 = QVBoxLayout(self.login_stk1)

        self.login_stk2 = QWidget()
        self.verticalLayout_2 = QVBoxLayout(self.login_stk2)

        self.login_stk3 = QWidget()
        self.verticalLayout_3 = QVBoxLayout(self.login_stk3)

        
        self.id_insert=QLineEdit()
        self.id_label1 = QLabel("ID")
        self.id_label2 = QLabel("@3frame.com")
        self.id_label2.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.id_layout1 = QHBoxLayout()
        self.id_layout2 = QVBoxLayout()
        self.id_layout1.addWidget(self.id_label1,stretch=1)
        self.id_layout1.addWidget(self.id_insert, stretch=3)
        self.id_layout2.addLayout(self.id_layout1)
        self.id_layout2.addWidget(self.id_label2)

        self.pw_insert=QLineEdit()
        self.pw_label1 = QLabel("PW")
        self.pw_layout = QHBoxLayout()
        self.pw_insert.setEchoMode(QLineEdit.Password)
        self.pw_layout.addWidget(self.pw_label1, stretch=1)
        self.pw_layout.addWidget(self.pw_insert, stretch=3)

        self.verticalLayout_1.addLayout(self.id_layout1,stretch=2)
        self.verticalLayout_1.addLayout(self.id_layout2,stretch=3)
        self.verticalLayout_1.addLayout(self.pw_layout,stretch=2)

        self.login_btn = QPushButton("Login")
        self.login_btn.clicked.connect(self.login_ck)
        self.verticalLayout_1.addWidget(self.login_btn,stretch=2)
        

        self.stk_w_login.addWidget(self.login_stk1)

        

        self.main_layout.addWidget(self.stk_w_login)
        
        self.setLayout(self.main_layout)
        self.id = self.id_insert.text()
        self.pw = self.pw_insert.text()
        #find = chrome_login(self.id,self.pw)
        #1page 내용

        self.date = QDate.currentDate()
        self.top_layout = QHBoxLayout()
        self.top_btn_layout = QVBoxLayout()
        self.ba_set_btn = QPushButton("설정")
        self.ba_btn = QPushButton("ID,PW 재입력")
        self.ba_label = QLabel("")
       # opacity_effect = QGraphicsOpacityEffect(self.ba_btn)
        #opacity_effect.setOpacity(0.5)
        #self.ba_btn.setGraphicsEffect(opacity_effect)
        self.top_btn_layout.addWidget(self.ba_set_btn,stretch=5)
        self.top_btn_layout.addWidget(self.ba_btn,stretch=5)
        
        self.top_layout.addWidget(self.ba_label,stretch=1)
        self.top_layout.addLayout(self.top_btn_layout,stretch=5)
        
        
        
        #self.ba_btn.setStyleSheet("background-color: ")
        self.s_btn = QPushButton("검색")

        self.name_layout = QHBoxLayout()
        self.name_label = QLabel("이름")
        self.name_insert=QLineEdit()
        self.name_layout.addWidget(self.name_label,stretch=2)
        self.name_layout.addWidget(self.name_insert,stretch=4)
        self.name_layout.addWidget(self.s_btn,stretch=2)

        self.class_label = QLabel("부서")
        self.class_insert=QtWidgets.QComboBox()
        self.class_layout = QHBoxLayout()
        self.class_layout.addWidget(self.class_label,stretch=2)
        self.class_layout.addWidget(self.class_insert,stretch=7)
        self.class_insert.currentTextChanged.connect(self.combo_search2)

        #self.verticalLayout_2.addLayout(self.top_layout,stretch=1)
        self.verticalLayout_2.addLayout(self.name_layout,stretch=2)
        self.verticalLayout_2.addLayout(self.class_layout,stretch=2)
        
        self.ba_btn.clicked.connect(self.ba_btn_ck)
        
        self.stk_w_login.addWidget(self.login_stk2)
        self.s_btn.clicked.connect(self.combo_search)
       


        self.inday_label = QLabel("입사일")
        self.inday_insert=QLabel()
        self.inday_insert.setAlignment(Qt.AlignRight)
        self.inday_layout = QHBoxLayout()
        self.inday_layout.addWidget(self.inday_label,stretch=2)
        self.inday_layout.addWidget(self.inday_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.inday_layout,stretch=2)

        self.rest_label = QLabel("발생연차")
        self.rest_insert=QLabel()
        self.rest_insert.setAlignment(Qt.AlignRight)
        self.rest_layout = QHBoxLayout()
        self.rest_layout.addWidget(self.rest_label,stretch=2)
        self.rest_layout.addWidget(self.rest_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.rest_layout,stretch=2)
        

        self.use_label = QLabel("사용연차")
        self.use_insert=QLabel()
        self.use_insert.setAlignment(Qt.AlignRight)
        self.use_layout = QHBoxLayout()
        self.use_layout.addWidget(self.use_label,stretch=2)
        self.use_layout.addWidget(self.use_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.use_layout,stretch=2)
        
        self.use2_label = QLabel("남은연차")
        self.use2_insert=QLabel()
        self.use2_insert.setAlignment(Qt.AlignRight)
        self.use2_layout = QHBoxLayout()
        self.use2_layout.addWidget(self.use2_label,stretch=2)
        self.use2_layout.addWidget(self.use2_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.use2_layout,stretch=2)
        
        self.send_label = QLabel("메일 발송일")
        self.send_insert=QLabel()
        self.send_insert.setAlignment(Qt.AlignRight)
        self.send_layout = QHBoxLayout()
        self.send_layout.addWidget(self.send_label,stretch=2)
        self.send_layout.addWidget(self.send_insert,stretch=7)
        self.verticalLayout_2.addLayout(self.send_layout,stretch=2)
        self.verticalLayout_2.addLayout(self.top_layout,stretch=1)

        self.db_up = Db_update(parent=self)
        self.db_up.label_msg.connect(self.label_ck)
        self.time_ck =time_ck(parent=self)
        self.time_ck.time_msg.connect(self.time_ck1)
        self.time_ck.label_msg.connect(self.label_ck)
        self.time_ck.send_msg.connect(self.send_ck)
        self.time_ck.start()
        self.send_m = send_mail(parent=self)


        self.time_h_insert=QtWidgets.QComboBox()
        self.time_m_insert=QtWidgets.QComboBox()
        self.time_layout = QHBoxLayout()
        self.verticalLayout_3.addWidget(self.login_btn,stretch=2)

    @pyqtSlot(str)
    def send_ck(self,msg) :
        
        if msg == "send":
            
            self.send_m.start()

    @pyqtSlot(str)
    def time_ck1(self,msg) :
        
        if msg == "start":
            
            self.db_up.start()
    
    @pyqtSlot(str)
    def label_ck(self,msg) :
        
        if msg == "DB":
            
            self.ba_label.setText("DB 업데이트 중")  

        if msg == "DB up":
            self.ba_label.setText("DB 업데이트 완료")  

        if msg =="send":
            print("확인")
            self.ba_label.setText("메일 발송 중")
        else :

            self.ba_label.setText(msg)

        
    def login_ck(self):
        if self.id_insert.text() == "" and self.pw_insert.text() == "":
            QMessageBox.information(self,"확인","ID,PW를 입력해주세요")
        elif self.id_insert.text() == "" :
            QMessageBox.information(self,"확인","ID를 입력해주세요")
        elif self.pw_insert.text() == "" :
            QMessageBox.information(self,"확인","PW를 입력해주세요")
        else :
            self.stk_w_login.setCurrentWidget(self.login_stk2)

    def ba_btn_ck(self) :
        self.stk_w_login.setCurrentWidget(self.login_stk1)
    def mi_btn_ck(self) :
        self.stk_w_login.setCurrentWidget(self.login_stk2)
    
   

    def combo_search2(self,item):
        try :
            name = self.name_insert.text()
            if name == "" :
                #QMessageBox.information(self,"확인","이름을 입력해주세요")
                self.clear_value()
            else :
                for row in database_ws.iter_rows(min_row=2):
                    if row[1].value == name :
                        if row[3].value == item:
                            in_d = row[4].value
                            r_d = row[5].value
                            u_d=row[6].value
                            l_d=row[7].value
                self.inday_insert.setText(str(in_d))
                self.rest_insert.setText(str(r_d))
                self.use_insert.setText(str(u_d)+"일")
                self.use2_insert.setText(str(l_d)+"일")
            
                name = self.name_insert.text()
                for row1 in database_ws2.iter_rows(min_row=2):
                    if row1[1].value == name :
                        if row1[3].value == item:
                            s_d = row1[4].value
                self.send_insert.setText(str(s_d))
        except OSError:
            pass


    def combo_search(self):
        try :
            name = self.name_insert.text()
            self.clearComboBoxItem()
            if name =="" :
                QMessageBox.information(self,"확인","이름을 입력해주세요")
                self.clear_value()
            else :    
                for row in database_ws.iter_rows(min_row=2):
                        
                    if row[1].value == name :
                    
                        self.h_class =row[3].value 
                        self.addComboBoxItem()
                
                    else :
                        
                        pass  
        except OSError :
            pass
    
    def clear_value(self) :
        self.inday_insert.setText("")
        self.rest_insert.setText("")
        self.use_insert.setText("")
        self.use2_insert.setText("")
        self.send_insert.setText("")
    def clearComboBoxItem(self) :
        self.class_insert.clear()
    
    def addComboBoxItem(self) :
        self.class_insert.addItem(self.h_class)
    
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = Main()
    main.show()
    sys.exit(app.exec_())
